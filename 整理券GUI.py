import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re

def sanitize_filename(s):
    return re.sub(r'[\\/*?:"<>|]', '_', s)

def create_tickets_excel(title, expiry, start_number, rows, cols, output_dir, font_name):
    ROWS_PER_TICKET = 3
    COLS_PER_TICKET = 2
    TICKETS_PER_ROW = cols
    TICKETS_PER_COL = rows
    START_NUMBER = start_number
    TICKET_TITLE = title
    EXPIRY_TEXT = expiry
    FILENAME_TPL = "{title}No.{start:03d}~.xlsx"

    font_title  = Font(name=font_name, bold=True, color="FFFFFF", size=12)
    font_number = Font(name=font_name, bold=True, size=18)
    font_date   = Font(name=font_name, size=8)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    fill_blue  = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
    fill_gray  = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = TICKET_TITLE

    for col in range(1, TICKETS_PER_ROW * COLS_PER_TICKET + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12
    for row in range(1, TICKETS_PER_COL * ROWS_PER_TICKET + 1):
        ws.row_dimensions[row].height = 20

    ticket_number = START_NUMBER
    for i in range(TICKETS_PER_COL):
        for j in range(TICKETS_PER_ROW):
            start_row = i * ROWS_PER_TICKET + 1
            start_col = j * COLS_PER_TICKET + 1
            end_row   = start_row + ROWS_PER_TICKET - 1
            end_col   = start_col + COLS_PER_TICKET - 1

            cell_title = ws.cell(row=start_row, column=start_col)
            ws.merge_cells(start_row=start_row, start_column=start_col,
                           end_row=start_row, end_column=end_col)
            cell_title.value = TICKET_TITLE
            cell_title.fill  = fill_blue
            cell_title.font  = font_title
            cell_title.alignment = align_center

            cell_num = ws.cell(row=start_row + 1, column=start_col)
            ws.merge_cells(start_row=start_row + 1, start_column=start_col,
                           end_row=start_row + 1, end_column=end_col)
            cell_num.value = f"No.{ticket_number:03d}"
            cell_num.font  = font_number
            cell_num.fill  = fill_white
            cell_num.alignment = align_center

            cell_date = ws.cell(row=start_row + 2, column=start_col)
            ws.merge_cells(start_row=start_row + 2, start_column=start_col,
                           end_row=start_row + 2, end_column=end_col)
            cell_date.value = EXPIRY_TEXT
            cell_date.fill  = fill_gray
            cell_date.font  = font_date
            cell_date.alignment = align_center

            for r in range(start_row, end_row + 1):
                for c in range(start_col, end_col + 1):
                    ws.cell(row=r, column=c).border = thin_border

            ticket_number += 1

    safe_title = sanitize_filename(TICKET_TITLE)
    filename = os.path.join(
        output_dir,
        FILENAME_TPL.format(title=safe_title, start=START_NUMBER)
    )
    wb.save(filename)

class TicketApp:
    def __init__(self, root):
        self.root = root
        self.root.title("整理券ジェネレーター Ver1.0.0")
        self.root.geometry("600x420")
        self.root.resizable(True, True)

        label_font = ("メイリオ", 12)
        entry_font = ("メイリオ", 12)

        padx = 15
        pady = 8

        self.root.grid_columnconfigure(0, weight=0)
        self.root.grid_columnconfigure(1, weight=1)

        for i in range(8):
            self.root.grid_rowconfigure(i, weight=0)
        self.root.grid_rowconfigure(7, weight=1)

        tk.Label(root, text="タイトル", font=label_font).grid(row=0, column=0, sticky="e", padx=padx, pady=pady)
        self.title_entry = tk.Entry(root, font=entry_font)
        self.title_entry.insert(0, "整理券")
        self.title_entry.grid(row=0, column=1, sticky="ew", padx=padx, pady=pady)

        tk.Label(root, text="有効期限", font=label_font).grid(row=1, column=0, sticky="e", padx=padx, pady=pady)
        self.expiry_entry = tk.Entry(root, font=entry_font)
        self.expiry_entry.insert(0, "XXXX年XX月XX日のみ有効")
        self.expiry_entry.grid(row=1, column=1, sticky="ew", padx=padx, pady=pady)

        tk.Label(root, text="開始番号", font=label_font).grid(row=2, column=0, sticky="e", padx=padx, pady=pady)
        self.start_num = tk.Entry(root, font=entry_font)
        self.start_num.insert(0, "1")
        self.start_num.grid(row=2, column=1, sticky="ew", padx=padx, pady=pady)

        tk.Label(root, text="行（縦）", font=label_font).grid(row=3, column=0, sticky="e", padx=padx, pady=pady)
        self.rows_entry = tk.Entry(root, font=entry_font)
        self.rows_entry.insert(0, "10")
        self.rows_entry.grid(row=3, column=1, sticky="ew", padx=padx, pady=pady)

        tk.Label(root, text="列（横）", font=label_font).grid(row=4, column=0, sticky="e", padx=padx, pady=pady)
        self.cols_entry = tk.Entry(root, font=entry_font)
        self.cols_entry.insert(0, "3")
        self.cols_entry.grid(row=4, column=1, sticky="ew", padx=padx, pady=pady)

        tk.Label(root, text="フォント", font=label_font).grid(row=5, column=0, sticky="e", padx=padx, pady=pady)
        self.font_var = tk.StringVar()
        font_options = ["メイリオ", "ヒラギノ丸ゴ ProN W4", "Arial", "ＭＳ ゴシック", "游ゴシック"]
        self.font_combo = ttk.Combobox(root, textvariable=self.font_var, values=font_options, font=entry_font, state="readonly")
        self.font_combo.current(0)
        self.font_combo.grid(row=5, column=1, sticky="ew", padx=padx, pady=pady)

        tk.Button(root, text="出力フォルダ選択", font=label_font, command=self.choose_dir).grid(row=6, column=0, sticky="ew", padx=padx, pady=pady)
        self.output_dir = tk.Label(root, text="未選択", font=label_font, anchor="w")
        self.output_dir.grid(row=6, column=1, sticky="ew", padx=padx, pady=pady)

        tk.Button(root, text="生成", font=("メイリオ", 14, "bold"), bg="#4A90E2", fg="white", command=self.generate).grid(row=7, column=0, columnspan=2, sticky="ew", padx=padx, pady=(20, 30))

    def choose_dir(self):
        path = filedialog.askdirectory()
        if path:
            self.output_dir.config(text=path)

    def generate(self):
        try:
            title = self.title_entry.get()
            expiry = self.expiry_entry.get()
            start_number = int(self.start_num.get())
            rows = int(self.rows_entry.get())
            cols = int(self.cols_entry.get())
            output = self.output_dir.cget("text")
            font_name = self.font_var.get()
            if output == "未選択":
                raise ValueError("出力フォルダが未選択です。")

            create_tickets_excel(title, expiry, start_number, rows, cols, output, font_name)
            messagebox.showinfo("成功", "整理券を生成しました。")

        except Exception as e:
            messagebox.showerror("エラー", str(e))

if __name__ == "__main__":
    root = tk.Tk()
    app = TicketApp(root)
    root.mainloop()