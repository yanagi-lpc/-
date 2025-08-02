from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re

# ====================  定数設定  ====================
ROWS_PER_TICKET = 3
COLS_PER_TICKET = 2

TICKETS_PER_ROW = 3
TICKETS_PER_COL = 10
START_NUMBER = 1

TICKET_TITLE = "整理券"
EXPIRY_TEXT = "XXXX年XX月XX日のみ有効"
FILENAME_TPL = "【{title}】No.{start:03d}~.xlsx"
OUTPUT_DIR = None

# フォント設定
FONT_NAME = "ヒラギノ丸ゴ ProN W4"   # 例: "Arial", "ＭＳ ゴシック", "Calibri"
FONT_SIZE_TITLE = 12
FONT_SIZE_NUMBER = 18
FONT_SIZE_DATE = 8
FONT_COLOR_TITLE = "FFFFFF"

# ---------------------------------------------

def sanitize_filename(s):
    return re.sub(r'[\\/*?:"<>|]', '_', s)

# ======= スタイル定義 =======
font_title = Font(name=FONT_NAME, bold=True, color=FONT_COLOR_TITLE, size=FONT_SIZE_TITLE)
font_number = Font(name=FONT_NAME, bold=True, size=FONT_SIZE_NUMBER)
font_date = Font(name=FONT_NAME, size=FONT_SIZE_DATE)
align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

fill_blue = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
fill_gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000'),
)

# ========= ワークブック作成 =========
wb = Workbook()
ws = wb.active
ws.title = TICKET_TITLE

for col in range(1, TICKETS_PER_ROW * COLS_PER_TICKET + 1):
    ws.column_dimensions[get_column_letter(col)].width = 12
for row in range(1, TICKETS_PER_COL * ROWS_PER_TICKET + 1):
    ws.row_dimensions[row].height = 20

ticket_number = START_NUMBER
for i in range(TICKETS_PER_COL):
    for j in range(TICKETS_PER_ROW):
        start_row = i * ROWS_PER_TICKET + 1
        start_col = j * COLS_PER_TICKET + 1
        end_row = start_row + ROWS_PER_TICKET - 1
        end_col = start_col + COLS_PER_TICKET - 1

        # タイトル
        cell_title = ws.cell(row=start_row, column=start_col)
        ws.merge_cells(start_row=start_row, start_column=start_col,
                       end_row=start_row, end_column=end_col)
        cell_title.value = TICKET_TITLE
        cell_title.fill = fill_blue
        cell_title.font = font_title
        cell_title.alignment = align_center

        # 整理券番号
        cell_num = ws.cell(row=start_row + 1, column=start_col)
        ws.merge_cells(start_row=start_row + 1, start_column=start_col,
                       end_row=start_row + 1, end_column=end_col)
        cell_num.value = f"No.{ticket_number:03d}"
        cell_num.font = font_number
        cell_num.fill = fill_white
        cell_num.alignment = align_center

        # 有効期限
        cell_date = ws.cell(row=start_row + 2, column=start_col)
        ws.merge_cells(start_row=start_row + 2, start_column=start_col,
                       end_row=start_row + 2, end_column=end_col)
        cell_date.value = EXPIRY_TEXT
        cell_date.fill = fill_gray
        cell_date.font = font_date
        cell_date.alignment = align_center

        # 外枠罫線
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                ws.cell(row=r, column=c).border = thin_border

        ticket_number += 1

if OUTPUT_DIR is None:
    OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

safe_title = sanitize_filename(TICKET_TITLE)
filename = os.path.join(
    OUTPUT_DIR,
    FILENAME_TPL.format(title=safe_title, start=START_NUMBER)
)

wb.save(filename)